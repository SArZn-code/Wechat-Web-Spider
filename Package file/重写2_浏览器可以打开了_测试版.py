from bs4 import BeautifulSoup as BS
from requests import post
from requests import get
from json import load
from datetime import timedelta
from datetime import date
import pandas as pd
from time import sleep
from sys import exit
import os
from warnings import filterwarnings
import openpyxl as op
import winreg
from pathlib import Path
import keyboard

# 重构后的文件agent依赖于txt文件, 整合1和2之后的文件则不需要agent.txt文件
class A:

    def __init__(self) -> None: # 公用??
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
        path_1 = winreg.QueryValueEx(key, "Desktop")[0]
        self.desktop_path = Path(path_1.replace("\\","/"))
        with open(f"{self.desktop_path / 'web spider' / 'agent.txt'}","r",encoding='utf-8') as fin:
            self.agent = fin.readlines()[0]
            
        self.index_sheet_2_num = 0
        self.list_index_1000 = []
        self.account_1000 = []
        self.title_1000 = []
        self.time_1000 = []
        self.read_num_1000 = []
        self.old_like_num_1000 = []

        self.category = []
        self.list_index_s1 = []
        self.account_s1 = []
        self.publish_date_num_s1 = []
        self.publish_num_s1 = []
        self.read_num_s1 = []
        self.first_read_num_s1 = []
        self.old_like_num_s1 = []

        self.py_2_biz_with_url = {}

        self.all_data = open(f"{self.desktop_path / 'web spider' / 'all_data.txt'}","w",encoding="utf-8")

    def start(self): #公用
        print("\n\n=========================================*^_^*微信公众号文章参数采集*^_^*=============================================")
        print("请确定是否要执行该程序\n输入yes或no:", end="")
        while(True):
            bl_2 = input().lower()
            if(bl_2 == "yes"):
                break
            elif(bl_2 == "no"):
                exit()
            else:
                print("无效输入，请重新确认是否要执行该程序:", end="")
        print("\n==================================================================================================================\n")
    def time_restrict(self): #公用
        today_date = date.today()
        today_week = today_date.strftime("%A")
        dict_week = {
            "Monday":5,
            "Tuesday":6,
            "Wednesday":0,
            "Thursday":1,
            "Friday":2,
            "Saturday":3,
            "Sunday":4,
        }
        self.this_Wednesday = today_date - timedelta(days=dict_week[today_week])
        self.last_Thrusday = self.this_Wednesday - timedelta(days=6)

        print("如果非此时间段,请通过左右键来切换日期, 成功后Enter")
        print(self.last_Thrusday,self.this_Wednesday, end='\r')

        def on_key_event(event):
            if event.name == 'left':
                self.this_Wednesday = self.this_Wednesday - timedelta(days=7)
                self.last_Thrusday = self.last_Thrusday - timedelta(days=7)
                print(self.last_Thrusday,self.this_Wednesday, end='\r')

            elif event.name == 'right':
                self.this_Wednesday = self.this_Wednesday + timedelta(days=7)
                self.last_Thrusday = self.last_Thrusday + timedelta(days=7)
                print(self.last_Thrusday,self.this_Wednesday, end='\r')

        keyboard.on_press(on_key_event)
        keyboard.wait('enter')


    def first_check_file_and_dir(self):
        # py_2 = self.desktop_path / 'web spider' / 'py_2 need url.txt'
        # if py_2.exists():
        #     os.remove(py_2)
        
        for sheet in [1,2]:
            file_path = self.desktop_path / 'web spider' / f'sheet {sheet}_{self.last_Thrusday}--{self.this_Wednesday}.xlsx'
            if file_path.exists():
                os.remove(file_path)

    def check(self):
        self.files = []
        # account = ""
        for root, dirs, temp_files in os.walk(self.desktop_path / 'web spider' / 'all of url'):
            self.files = temp_files
    #     print("\n请务必确保每一个公众号都打开并完整刷新\n")
    #     for acco in self.files:
    #         print(acco.split('_')[0])
    #     # input("\n!!完成后再按Enter\n")
    #     os.system('pause')

    # def fiddler_get(self):
    #     f_allurl = open(f"{self.desktop_path / 'web spider' / 'py_2 need url.txt'}","r",encoding="utf-8")
    #     allurl = f_allurl.readlines()
    #     f_allurl.close()

    #     for each in allurl:
    #         each_url = each.strip("\n")
    #         # print(each_url)
    #         each_biz = each_url.split("__biz=")[2].split("&")[0][0:-6] + "=="
    #         # print(each_biz)
    #         self.py_2_biz_with_url[each_biz] = each_url

    def backbone(self):
        fakeid_biz = {}
        with open(f"{self.desktop_path / 'web spider' / '数据_公众号信息.json'}","r",encoding="utf-8") as f_fakeid_biz:
            fakeid_biz = load(f_fakeid_biz)

        surplus = len(self.files)

        for i in self.files:
            account = i.split("_")[0]
            surplus -= 1
            print(f"{account}开始收集,还剩{surplus}个公众号")

            
            list_info = []
            open_mid_with_title = {}
            open_title_with_time = {}
            with open(f"{self.desktop_path / 'web spider' / 'all of url' / f'{account}_url.txt'}","r",encoding="utf-8") as open_url:
                list_info = open_url.readlines()
            with open(f"{self.desktop_path / 'web spider' / 'mid_with_title' / f'{account}.json'}","r") as open_mid:
                open_mid_with_title = load(open_mid)
            with open(f"{self.desktop_path / 'web spider' / 'title_with_time' / f'{account}.json'}","r") as open_title:
                open_title_with_time = load(open_title)
            #后续bug_check
            if(len(list_info) != len(open_mid_with_title)):
                for i in range(0,10):
                    print('sheet 2标题出现错误, 请结束后修改')
            if(len(list_info) != len(open_title_with_time)):
                for i in range(0,10):
                    print('sheet 2时间出现错误, 请结束后修改')
            ############
            self.account_s1.append(account)
            seg_publish_num_s1 = len(list_info)
            seg_read_num_s1 = 0     
            seg_old_like_num_s1 = 0 
            dict_time = []
            seg_publish_date_num_s1 = 0 
            seg_first_read_num_s1 = 0

            print(f"共{len(list_info)}篇")

            # 优化
            # temp = fakeid_biz[account]
            # temp_bl = 1
            # while temp_bl:
            #     try: 
            #         url = "https://" + self.py_2_biz_with_url[temp]
            #         print(url)
            #         temp_bl = 0
            #     except:
            #         input(f'{account}没有获取到, 请重新点击, 点完后按Enter')
            #         self.fiddler_get()
            

            # 优化处理
            # change_mid = ''
            # ii = 0

            for i in list_info:
                mid = i.split("&")[1].split("=")[1]
                idx = i.split("&")[2].split("=")[1]
                sn = i.split("&")[3].split("=")[1]
                _biz = i.split("&")[0].split("?")[1].split("_biz=")[1]

                # 新加的 , 直接在浏览器中打开公众号
                url = i

                #优化处理
                # if(mid == change_mid):
                #     pass
                # else:
                #     ii = 0
                # ii += 1
                mid = mid + f'_{idx}'

                headers = {
                    # "User-Agent": str(self.agent),
                    "User-Agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 NetType/WIFI MicroMessenger/7.0.20.1781(0x6700143B) WindowsWechat(0x63090b13) XWEB/9185',
                }
                data = {
                    "mid": mid,
                    "idx": idx,
                    "sn": sn,
                    "__biz": _biz,
                    "is_only_read": "1",
                    "is_temp_url": "0",
                    "appmsg_type": "9",
                    "reward_uin_count":"0",
                }
                
                result = get(url,headers=headers)
                print(result.text)
                try:
                    secret = True
                    read_num = result.json()["appmsgstat"]["read_num"]
                    old_like_num = result.json()["appmsgstat"]["old_like_num"]
                    # 优化
                    seg_read_num_s1 += read_num
                    seg_old_like_num_s1 += old_like_num

                except KeyError:
                    # secret
                    secret = False
                    # 优化
                    read_num = -1
                    old_like_num = -1
                    # 优化
                    seg_read_num_s1 = -1
                    seg_old_like_num_s1 = -1

                    print("请求过于频繁,read_num & old_like_num 用 -1 取代")
                    # # 优化
                    # suspend_time = 5000
                    # for suspend in range(suspend_time, -1, -1):
                    #     sleep(1)
                    #     print("等待中...{:>4}秒".format(suspend), end="\r")
                    # print("\r")


                mid_title = open_mid_with_title[mid]
                # 优化
                new_title = mid + '|' + mid_title
                title_time = open_title_with_time[new_title][5:]

                if(title_time in dict_time):
                    pass
                else:
                    dict_time.append(title_time)
                    seg_publish_date_num_s1 += 1
                    # 优化
                    if not secret:
                        seg_first_read_num_s1 = -1
                    else:
                        seg_first_read_num_s1 += read_num

                self.all_data.write(f"{account}:--:{mid_title}:--:{title_time}:--:{read_num}:--:{old_like_num}\n")

                if(read_num >= 1000):
                    self.index_sheet_2_num += 1
                    self.account_1000.append(account)
                    self.title_1000.append(mid_title)
                    self.read_num_1000.append(read_num)
                    self.old_like_num_1000.append(old_like_num)
                    title_time = round(int(title_time.split("-")[0]) + int(title_time.split("-")[1])/100, 2)
                    self.time_1000.append(title_time)

                sleep_time = 5 
                for remaining_time in range(sleep_time, -1, -1):
                    sleep(1)
                    print(f"等待爬取下一篇文章中...{remaining_time}秒", end="\r")
                print("\r")

            self.publish_num_s1.append(seg_publish_num_s1)
            self.read_num_s1.append(seg_read_num_s1)# 优化
            self.old_like_num_s1.append(seg_old_like_num_s1)# 优化
            self.first_read_num_s1.append(seg_first_read_num_s1)# 优化
            self.publish_date_num_s1.append(seg_publish_date_num_s1)

            sleep_time = 10 
            for remaining_time in range(sleep_time, -1, -1):
                sleep(1)
                print("等待爬取下一公众号中...{:>2}秒".format(remaining_time), end="\r")
            print("\r")
            print("\n")
            
        self.all_data.close()

    def write_in(self):
        file_path = []
        for sheet in [1,2]:
            file_path.append(f"{self.desktop_path / 'web spider' / f'sheet {sheet}_{self.last_Thrusday}--{self.this_Wednesday}.xlsx'}")
        
        for index_sheet_2 in range(1,self.index_sheet_2_num+1):
            self.list_index_1000.append(index_sheet_2)

        pd_account_s2 = pd.Series(self.account_1000, index=self.list_index_1000, name="公众号")
        pd_title_s2 = pd.Series(self.title_1000, index=self.list_index_1000, name="文章")
        pd_time_s2 = pd.Series(self.time_1000, index=self.list_index_1000, name="发布日期")
        pd_read_num_s2 = pd.Series(self.read_num_1000, index=self.list_index_1000, name="阅读量")
        pd_old_like_num_s2 = pd.Series(self.old_like_num_1000, index=self.list_index_1000, name="点赞量")
        pd_dict_s2 = {
            pd_account_s2.name: pd_account_s2,
            pd_title_s2.name: pd_title_s2,
            pd_time_s2.name: pd_time_s2,
            pd_read_num_s2.name: pd_read_num_s2,
            pd_old_like_num_s2.name: pd_old_like_num_s2,
        }
        df = pd.DataFrame(pd_dict_s2)
        df = df.set_index("公众号")
        df.to_excel(file_path[1])

        with open(f"{self.desktop_path / 'web spider' / '数据_公众号and类别.json'}","r",encoding="utf-8") as file:
            all_account = load(file)

        n = 0
        for acc in list(all_account.keys()):
            if acc not in self.account_s1:
                self.account_s1.append(acc)
                n += 1

        for acc in self.account_s1:
            self.category.append(all_account[acc])

        for index_sheet_1 in range(1,len(self.account_s1)+1):
            self.list_index_s1.append(index_sheet_1)

        self.publish_date_num_s1 += [0]*n
        self.publish_num_s1 += [0]*n
        self.read_num_s1 += [0]*n
        self.first_read_num_s1 += [0]*n
        self.old_like_num_s1 += [0]*n

        pd_account_s1 = pd.Series(self.account_s1, index=self.list_index_s1, name="公众号")
        pd_date_num_s1 = pd.Series(self.publish_date_num_s1, index=self.list_index_s1, name="发布次数").fillna(0)
        pd_publish_num_s1 = pd.Series(self.publish_num_s1, index=self.list_index_s1, name="发布篇数").fillna(0)
        pd_read_num_s1 = pd.Series(self.read_num_s1, index=self.list_index_s1, name="阅读总量").fillna(0)
        pd_first_read_num_s1 = pd.Series(self.first_read_num_s1, index=self.list_index_s1, name="头条阅读量").fillna(0)
        pd_old_like_num_s1 = pd.Series(self.old_like_num_s1, index=self.list_index_s1, name="点赞量").fillna(0)
        pd_category = pd.Series(self.category, index=self.list_index_s1, name="类别")
        pd_dict_s1 = {
            pd_account_s1.name: pd_account_s1,
            pd_date_num_s1.name: pd_date_num_s1,
            pd_publish_num_s1.name: pd_publish_num_s1,
            pd_read_num_s1.name: pd_read_num_s1,
            pd_first_read_num_s1.name: pd_first_read_num_s1,
            pd_old_like_num_s1.name: pd_old_like_num_s1,
            pd_category.name: pd_category,
        }
        df = pd.DataFrame(pd_dict_s1)
        df = df.set_index("公众号")
        df.to_excel(file_path[0])

        background_color = ""
        word_style = '华文细黑'
        size = 11

        for son_file in file_path:
            workbook = op.load_workbook(son_file)

            sheet = workbook.active
            header_row = sheet[1]
            for cell in header_row:
                if(son_file == file_path[0]):
                    background_color = "F7CAAC"
                else:
                    background_color = "A8D08E"
                cell.fill = op.styles.PatternFill(fill_type = 'solid', fgColor = background_color)

            font = op.styles.Font(name = word_style, size = size, color="404040")
            alignment_center = op.styles.Alignment(horizontal="center", vertical="center")
            alignment_left = op.styles.Alignment(horizontal="left", vertical="center")
            
            if(son_file == file_path[0]):
                sheet_n = 1
            else:
                sheet_n = 2

            col_n = 0
            for col in sheet.iter_cols():
                row_n = 0
                for cell in col:
                
                    if(col_n == 0 and row_n != 0):
                        cell.alignment = alignment_left
                    else:
                        if(sheet_n == 2 and col_n == 1 and row_n != 0):
                            cell.alignment = alignment_left
                        else:
                            cell.alignment = alignment_center
            
                    cell.font = font

                    row_n += 1
                col_n += 1
            
            workbook.save(son_file)
            workbook.close()

#----------------------------------------------------------------------
def main():
    filterwarnings("ignore")
    #函数调用
    create_2 = A()
    create_2.start()
    create_2.time_restrict()
    create_2.first_check_file_and_dir()
    create_2.check()
    # create_2.fiddler_get()
    create_2.backbone()
    create_2.write_in()

    os.system("pause")
    
if __name__=='__main__':
    main()