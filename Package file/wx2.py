from requests import post
from json import load
from datetime import timedelta
from datetime import date
import pandas as pd
from time import sleep
from sys import exit
import os
from warnings import filterwarnings
import openpyxl as op
import winreg

filterwarnings("ignore")

today_date = date.today()
today_week = today_date.strftime("%A")
dict_week = {
    "Monday":5,
    "Tuesday":6,
    "Wednesday":0,
    "Thursday":1,
    "Friday":2,
    "Saturday":3,
    "Sunday":4,
}
this_Wednesday = today_date - timedelta(days=dict_week[today_week])
last_Thrusday = this_Wednesday - timedelta(days=6)

print("\n\n=========================================*^_^*微信公众号文章参数采集*^_^*=============================================")

print("请确定是否要执行该程序\n输入yes或no:", end="")
while(True):
    bl_2 = input().lower()
    if(bl_2 == "yes"):
        break
    elif(bl_2 == "no"):

        exit()
    else:
        print("无效输入，请重新确认是否要执行该程序:", end="")
print("\n==================================================================================================================")

print(last_Thrusday,this_Wednesday)
INPUT = input("请确认:Enter ; 如果非此时间段, 请输入对应时间段, 格式:xxxx-xx-xx;xxxx-xx-xx \n")
if INPUT == '':
    pass
else:
    last_Thrusday, this_Wednesday = INPUT.split(";")

key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
path_1 = winreg.QueryValueEx(key, "Desktop")[0]
path = path_1.replace("\\","/")
# 上面公用

index_sheet_2_num = 0
list_index_1000 = []
account_1000 = []
title_1000 = []
time_1000 = []
read_num_1000 = []
old_like_num_1000 = []

category = []
list_index_s1 = []
account_s1 = []
publish_date_num_s1 = []
publish_num_s1 = []
read_num_s1 = []
first_read_num_s1 = []
old_like_num_s1 = []

if os.path.exists(f"{path}/web spider/py_2 need url.txt"):
    os.remove(f"{path}/web spider/py_2 need url.txt")

all_data = open(f"{path}/web spider/all_data.txt","w",encoding="utf-8")

with open(f"{path}/web spider/agent.txt","r",encoding='utf-8') as fin:
    agent = fin.readlines()[0]

files = []
account = ""
for root, dirs, temp_files in os.walk(f"{path}/web spider/all of url"):
    files = temp_files
print("请务必确保每一个公众号都打开并完整刷新")
for acco in files:
    print(acco.split('_')[0])
input("!!完成后再按Enter")

py_2_biz_with_url = {}
with open(f"{path}/web spider/py_2 need url.txt","r",encoding="utf-8") as f_allurl:
    allurl = f_allurl.readlines()
for each in allurl:
    each_url = each.strip("\n")
    each_biz = each_url.split("__biz=")[1].split("&")[0][0:-6] + "==" 
    py_2_biz_with_url[each_biz] = each_url

fakeid_biz = {}
with open(f"{path}/web spider/数据_公众号信息.json","r",encoding="utf-8") as f_fakeid_biz:
    fakeid_biz = load(f_fakeid_biz)

surplus = len(files)

for i in files:
    account = i.split("_")[0]
    surplus -= 1
    print(f"{account}开始收集,还剩{surplus}个公众号")


    list_info = []
    open_mid_with_title = {}
    open_title_with_time = {}
    with open(f"{path}/web spider/all of url/{account}_url.txt","r",encoding="utf-8") as open_url:
        list_info = open_url.readlines()
    # change_mid = ''
    # ii = 1
    # for new_mid in range(0, len(list_info)):
    #     mid = list_info[new_mid].split("&")[1].split("=")[1]
    #     if mid == change_mid:
    #         ii += 1
    #         index = list_info[new_mid].replace(mid, mid + f'_{ii}')
    #         index += len(mid)
    #     else:
    #         change_mid = mid
    #         ii = 1
    #         index = list_info[new_mid].index(mid)
    #         index += len(mid)
    with open(f"{path}/web spider/mid_with_title/{account}.json","r") as open_mid:
        open_mid_with_title = load(open_mid)
    with open(f"{path}/web spider/title_with_time/{account}.json","r") as open_title:
        open_title_with_time = load(open_title)
        
    if(len(list_info) != len(open_mid_with_title)):
        for i in range(0,10):
            print('sheet 2标题出现错误, 请结束后修改')
    if(len(list_info) != len(open_title_with_time)):
        for i in range(0,10):
            print('sheet 2时间出现错误, 请结束后修改')
    
    account_s1.append(account)
    seg_publish_num_s1 = len(list_info)
    seg_read_num_s1 = 0     
    seg_old_like_num_s1 = 0 
    dict_time = []
    seg_publish_date_num_s1 = 0 
    seg_first_read_num_s1 = 0  

    print(f"共{len(list_info)}篇")
    #优化
    temp = fakeid_biz[account]
    temp_bl = 1
    while temp_bl:
        try: 
            url = "https://" + py_2_biz_with_url[temp]
            temp_bl = 0
        except:
            input(f'{account}没有获取到, 请重新点击, 点完后按Enter')
            with open(f"{path}/web spider/py_2 need url.txt","r",encoding="utf-8") as f_allurl:
                allurl = f_allurl.readlines()
            for each in allurl:
                each_url = each.strip("\n")
                each_biz = each_url.split("__biz=")[1].split("&")[0][0:-6] + "==" 
                py_2_biz_with_url[each_biz] = each_url
        
    # 优化处理  
    # change_mid = ''
    # ii = 0

    for i in list_info:
        mid = i.split("&")[1].split("=")[1]
        idx = i.split("&")[2].split("=")[1]
        sn = i.split("&")[3].split("=")[1]
        _biz = i.split("&")[0].split("?")[1].split("_biz=")[1]

        #优化处理
        # if(mid != change_mid):
        #     ii = 0
        # ii += 1
        mid = mid + f'_{idx}'

        headers = {
            "User-Agent": agent,
        }
        data = {
            "mid": mid,
            "idx": idx,
            "sn": sn,
            "__biz": _biz,
            "is_only_read": "1",
            "is_temp_url": "0",
            "appmsg_type": "9",
            "reward_uin_count":"0",
        }

        result = post(url,headers=headers,data=data,verify=False)
        try:
            read_num = result.json()["appmsgstat"]["read_num"]
        except KeyError:
            print("请求过于频繁,等待继续")
            suspend_time = 5000
            for suspend in range(suspend_time, -1, -1):
                sleep(1)
                print("等待中...{:>4}秒".format(suspend), end="\r")
            print("\r")

        read_num = result.json()["appmsgstat"]["read_num"]
        old_like_num = result.json()["appmsgstat"]["old_like_num"]
        mid_title = open_mid_with_title[mid]
        # 优化
        new_title = mid + '|' + mid_title
        title_time = open_title_with_time[new_title][5:]

        seg_read_num_s1 += read_num
        seg_old_like_num_s1 += old_like_num
        if(title_time in dict_time):
            pass
        else:
            dict_time.append(title_time)
            seg_publish_date_num_s1 += 1
            seg_first_read_num_s1 += read_num

        all_data.write(f"{account}:--:{mid_title}:--:{title_time}:--:{read_num}:--:{old_like_num}\n")

        if(read_num >= 1000):
            index_sheet_2_num += 1
            account_1000.append(account)
            title_1000.append(mid_title)
            read_num_1000.append(read_num)
            old_like_num_1000.append(old_like_num)
            title_time = round(int(title_time.split("-")[0]) + int(title_time.split("-")[1])/100, 2)
            time_1000.append(title_time)

        sleep_time = 5 
        for remaining_time in range(sleep_time, -1, -1):
            sleep(1)
            print(f"等待爬取下一篇文章中...{remaining_time}秒", end="\r")
        print("\r")

    publish_num_s1.append(seg_publish_num_s1)
    read_num_s1.append(seg_read_num_s1)
    old_like_num_s1.append(seg_old_like_num_s1)
    first_read_num_s1.append(seg_first_read_num_s1)
    publish_date_num_s1.append(seg_publish_date_num_s1)

    print("\n")

    sleep_time = 10 
    for remaining_time in range(sleep_time, -1, -1):
        sleep(1)
        print("等待爬取下一公众号中...{:>2}秒".format(remaining_time), end="\r")
    print("\r")

all_data.close()


for file_path in [f"{path}/web spider/sheet 2_{last_Thrusday}--{this_Wednesday}.xlsx",f"{path}/web spider/sheet 1_{last_Thrusday}--{this_Wednesday}.xlsx"]:
    if os.path.exists(file_path):
        os.remove(file_path)

for index_sheet_2 in range(1,index_sheet_2_num+1):
    list_index_1000.append(index_sheet_2)

pd_account_s2 = pd.Series(account_1000, index=list_index_1000, name="公众号")
pd_title_s2 = pd.Series(title_1000, index=list_index_1000, name="文章")
pd_time_s2 = pd.Series(time_1000, index=list_index_1000, name="发布日期")
pd_read_num_s2 = pd.Series(read_num_1000, index=list_index_1000, name="阅读量")
pd_old_like_num_s2 = pd.Series(old_like_num_1000, index=list_index_1000, name="点赞量")
pd_dict_s2 = {
    pd_account_s2.name: pd_account_s2,
    pd_title_s2.name: pd_title_s2,
    pd_time_s2.name: pd_time_s2,
    pd_read_num_s2.name: pd_read_num_s2,
    pd_old_like_num_s2.name: pd_old_like_num_s2,
}
df = pd.DataFrame(pd_dict_s2)
df = df.set_index("公众号")
df.to_excel(f"{path}/web spider/sheet 2_{last_Thrusday}--{this_Wednesday}.xlsx")

with open(f"{path}/web spider/数据_公众号and类别.json","r",encoding="utf-8") as file:
    all_account = load(file)

n = 0
for acc in list(all_account.keys()):
    if acc not in account_s1:
        account_s1.append(acc)
        n += 1

for acc in account_s1:
    category.append(all_account[acc])

for index_sheet_1 in range(1,len(account_s1)+1):
    list_index_s1.append(index_sheet_1)

publish_date_num_s1 += [0]*n
publish_num_s1 += [0]*n
read_num_s1 += [0]*n
first_read_num_s1 += [0]*n
old_like_num_s1 += [0]*n

pd_account_s1 = pd.Series(account_s1, index=list_index_s1, name="公众号")
pd_date_num_s1 = pd.Series(publish_date_num_s1, index=list_index_s1, name="发布次数").fillna(0)
pd_publish_num_s1 = pd.Series(publish_num_s1, index=list_index_s1, name="发布篇数").fillna(0)
pd_read_num_s1 = pd.Series(read_num_s1, index=list_index_s1, name="阅读总量").fillna(0)
pd_first_read_num_s1 = pd.Series(first_read_num_s1, index=list_index_s1, name="头条阅读量").fillna(0)
pd_old_like_num_s1 = pd.Series(old_like_num_s1, index=list_index_s1, name="点赞量").fillna(0)
pd_category = pd.Series(category, index=list_index_s1, name="类别")
pd_dict_s1 = {
    pd_account_s1.name: pd_account_s1,
    pd_date_num_s1.name: pd_date_num_s1,
    pd_publish_num_s1.name: pd_publish_num_s1,
    pd_read_num_s1.name: pd_read_num_s1,
    pd_first_read_num_s1.name: pd_first_read_num_s1,
    pd_old_like_num_s1.name: pd_old_like_num_s1,
    pd_category.name: pd_category,
}
df = pd.DataFrame(pd_dict_s1)
df = df.set_index("公众号")
df.to_excel(f"{path}/web spider/sheet 1_{last_Thrusday}--{this_Wednesday}.xlsx")

excel = [f"sheet 1_{last_Thrusday}--{this_Wednesday}",f"sheet 2_{last_Thrusday}--{this_Wednesday}"]
background_color = ""
word_style = '华文细黑'
size = 11

for i in excel:
    workbook = op.load_workbook(f'{path}/web spider/{i}.xlsx')

    sheet = workbook.active
    header_row = sheet[1]
    for cell in header_row:
        if(i == excel[0]):
            background_color = "F7CAAC"
        else:
            background_color = "A8D08E"
        cell.fill = op.styles.PatternFill(fill_type = 'solid', fgColor = background_color)

    font = op.styles.Font(name = word_style, size = size, color="404040")
    alignment_center = op.styles.Alignment(horizontal="center", vertical="center")
    alignment_left = op.styles.Alignment(horizontal="left", vertical="center")
    
    if(i == excel[0]):
        sheet_n = 1
    else:
        sheet_n = 2

    col_n = 0
    for col in sheet.iter_cols():
        row_n = 0
        for cell in col:
        
            if(col_n == 0 and row_n != 0):
                cell.alignment = alignment_left
            else:
                if(sheet_n == 2 and col_n == 1 and row_n != 0):
                    cell.alignment = alignment_left
                else:
                    cell.alignment = alignment_center
    
            cell.font = font

            row_n += 1
        col_n += 1
    
    workbook.save(f'{path}/web spider/{i}.xlsx')
    workbook.close()

os.system("pause")