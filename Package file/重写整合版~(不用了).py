# First
from requests import get
from time import sleep
from time import strftime
from time import localtime
from datetime import timedelta
from datetime import date
from sys import exit
from json import load
from json import loads
from json import JSONDecodeError
from json import dump
from warnings import filterwarnings
import os
from pathlib import Path
import winreg
from selenium.webdriver import EdgeOptions
from selenium.webdriver import Edge
from selenium.webdriver.edge.service import Service
# Second
from requests import post
import pandas as pd
import openpyxl as op

class Init:
    def __init__(self) -> None:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
        desktop_path_1 = winreg.QueryValueEx(key, "Desktop")[0]
        desktop_path_temp = desktop_path_1.replace("\\","/")
        self.desktop_path = Path(desktop_path_temp)
    def start(self): #公用
        print("\n\n=========================================*^_^*微信公众号文章参数采集*^_^*=============================================")
        print("请确定是否要执行该程序\n输入yes或no:", end="")
        while(True):
            bl_2 = input().lower()
            if(bl_2 == "yes"):
                break
            elif(bl_2 == "no"):
                exit()
            else:
                print("无效输入，请重新确认是否要执行该程序:", end="")
        print("\n==================================================================================================================\n")
    def time_restrict(self): #公用
        today_date = date.today()
        today_week = today_date.strftime("%A")
        dict_week = {
            "Monday":5,
            "Tuesday":6,
            "Wednesday":0,
            "Thursday":1,
            "Friday":2,
            "Saturday":3,
            "Sunday":4,
        }
        self.this_Wednesday = today_date - timedelta(days=dict_week[today_week])
        self.last_Thrusday = self.this_Wednesday - timedelta(days=6)

        return self.last_Thrusday, self.this_Wednesday

    
class Create_1:
    def __init__(self, desktop_path,last_Thrusday, this_Wednesday) -> None:
        self.desktop_path = desktop_path
        self.last_Thrusday = last_Thrusday
        self.this_Wednesday = this_Wednesday

    def first_check_file_and_dir(self):
        path = self.desktop_path / 'web spider'
        list_file =["all of url","mid_with_title","title_with_time"]
        if not path.exists():
            path.mkdir()
        else:
            # 更新
            path_drive = ''
            # 检测如果存在多个, 该如何选择. 采用通配符模式
            path_1 = path / '数据_公众号and类别.json'
            path_2 = path / '数据_公众号信息.json'
            while not path_1.exists() and not path_2.exists():
                print('关键json文件未找到,请添加后继续程序')

            for file_i in list_file:
                path_i = path / file_i
                if not path_i.exists():
                    path_i.mkdir()
                else:
                    for son_file in os.listdir(path_i):
                        os.remove(path_i / son_file)

    def time_limit(self):
        with open(f"{self.desktop_path / 'web spider'/ '采集时间.txt'}","w+") as get_limit_date:
            old_content = get_limit_date.read()
            get_limit_date.seek(0)
            get_limit_date.write(str(self.last_Thrusday)+ "--" + str(self.this_Wednesday) + "\n" + old_content)
        print(self.last_Thrusday,self.this_Wednesday)
        input("请确认:Enter")


    def drive(self):
        url = "https://mp.weixin.qq.com"
        self.url_new = ''
        option = EdgeOptions()
        option.add_experimental_option('excludeSwitches', ['enable-automation'])
        option.add_experimental_option('useAutomationExtension', False)
        self.driver = Edge(service=Service(f"{self.desktop_path}/web spider/edge_driver.exe"),options=option)
        self.driver.get(url)
        #更新
        sleep(25)
        while(True):
            if(url == self.driver.current_url):
                sleep(1)
            else:
                self.url_new = self.driver.current_url
                break


    def get_url(self):

        self.token = ''
        self.cookies = ''
        self.agent = ''
        fragment = self.url_new.split('&')
        for i in fragment:
            if("token" in i):
                self.token = i.split("=")[1]
        for cookie in self.driver.get_cookies():
            self.cookies = self.cookies + cookie['name'] + '=' + cookie['value'] + ';'
        self.cookies = self.cookies.strip(';')

        file_agent = self.desktop_path / 'web spider' / 'agent.txt'

        self.agent = self.driver.execute_script("return navigator.userAgent;")

        self.driver.quit()

        with open(f"{self.desktop_path / 'web spider'/ '数据_公众号信息.json'}","r",encoding="utf-8") as file:
            self.dict_fakeid = load(file)
        
        return self.agent


    def backbone(self):

        for key in range(0,len(self.dict_fakeid)):
            account = list(self.dict_fakeid.keys())[key]

            sleep(3)

            print(f"{account},总计还剩{len(self.dict_fakeid)-key-1}个公众号")

            link_list = []
            dict_title_with_time = {}
            dict_mid_with_time = {}

            each_fakeid = self.dict_fakeid[account]

            page = 100 
            bool_date_limit = True

            for each_page in range(0,page):

                # url = f"https://mp.weixin.qq.com/cgi-bin/appmsg?action=list_ex&begin={each_page*5}&count=5&fakeid={each_fakeid}&type=9&query=&token={token}&lang=zh_CN&f=json&ajax=1"
                url = f"https://mp.weixin.qq.com/cgi-bin/appmsgpublish?sub=list&search_field=null&begin={each_page*5}&count=5&query=&fakeid={each_fakeid}&type=101_1&free_publish_type=1&sub_action=list_ex&token={self.token}&lang=zh_CN&f=json&ajax=1"
                header ={
                    'User-Agent': self.agent,
                    'Cookie':self.cookies,
                }
                
                result = get(url, headers=header, verify=False)
                receive_dict = result.json()
                def convert_to_dict(value):
                    try:
                        return loads(value)
                    except (JSONDecodeError, TypeError):
                        return value  
                processed_dict = {key: convert_to_dict(value) for key, value in receive_dict.items()}

                try:
                    processed_dict['publish_page']['publish_list']
                except KeyError:
                    print("请求过于频繁,等待继续. 若是本周初次执行,说明程序寄了")
                    suspend_time = 5000 
                    for suspend in range(suspend_time, -1, -1):
                        sleep(1)
                        print("等待中...{:>4}秒".format(suspend), end="\r")
                    print("\r")
                    processed_dict['publish_page']['publish_list']

                for i in processed_dict['publish_page']['publish_list']:
                    ii = {key: convert_to_dict(value) for key, value in i.items()}
                    for son_list in ii['publish_info']['appmsgex']:
                        local_time = localtime(int(son_list["create_time"]))
                        formal_time = strftime("%Y-%m-%d %H:%M:%S", local_time).split(" ")[0]

                        if(formal_time > str(self.this_Wednesday)):
                            continue

                        if(str(self.last_Thrusday) <= formal_time):
                            pass
                        else:
                            bool_date_limit = False

                        if(bool_date_limit == True):
                            pass
                        else:
                            break
                        
                        link = son_list["link"]
                        title = son_list["title"]
                        # mid = link.split("&")[1].split("=")[1]
                        mid = son_list["aid"]
                        link_list.append(link)
                        dict_mid_with_time[mid] = title 
                        dict_title_with_time[title] = formal_time

                        sleep_time = 5 
                        for remaining_time in range(sleep_time, -1, -1):
                            sleep(1)
                            print(f"等待爬取下一篇文章中...{remaining_time}秒", end="\r")
                        print("\r")
                    if(bool_date_limit == True):
                        pass
                    else:
                        break

                if(bool_date_limit == True):
                    pass
                else:
                    break
                sleep_time = 10 
                for remaining_time in range(sleep_time, -1, -1):
                    sleep(1)
                    print("等待爬取下一页中...{:>2}秒".format(remaining_time), end="\r")
                print("\r")

            self.write_in(link_list,account,dict_mid_with_time,dict_title_with_time)

    def write_in(self,link_list,account,dict_mid_with_time,dict_title_with_time):
        if(link_list == []):
            print(f"{account}本周无文章发表\n")
        else:
            print(f"{account}已全部爬取完毕,正在写入文件中...")

            with open(f"{self.desktop_path / 'web spider' / 'all of url' / f'{account}_url.txt'}","w",encoding="UTF-8") as file_url:
                for link in link_list:
                    file_url.write(link+"\n")
            print(f"{account}的链接读写完成;",end="")
            with open(f"{self.desktop_path / 'web spider' / 'mid_with_title' / f'{account}.json'}","w") as f_mid_json:
                dump(dict_mid_with_time,f_mid_json)
            with open(f"{self.desktop_path / 'web spider' / 'title_with_time' / f'{account}.json'}","w") as f_title_json:
                dump(dict_title_with_time,f_title_json)
            print("相关文件读写完成.已全部保存到桌面web spider文件夹中\n")
class Create_2:
    def __init__(self, desktop_path, agent,last_Thrusday, this_Wednesday) -> None:
        self.desktop_path = desktop_path
        self.agent = agent
        self.last_Thrusday = last_Thrusday
        self.this_Wednesday = this_Wednesday

        self.index_sheet_2_num = 0
        self.list_index_1000 = []
        self.account_1000 = []
        self.title_1000 = []
        self.time_1000 = []
        self.read_num_1000 = []
        self.old_like_num_1000 = []

        self.category = []
        self.list_index_s1 = []
        self.account_s1 = []
        self.publish_date_num_s1 = []
        self.publish_num_s1 = []
        self.read_num_s1 = []
        self.first_read_num_s1 = []
        self.old_like_num_s1 = []

        self.py_2_biz_with_url = {}

        self.all_data = open(f"{self.desktop_path / 'web spider' / 'all_data.txt'}","w",encoding="utf-8")

    def first_check_file_and_dir(self):
        py_2 = self.desktop_path / 'web spider' / 'py_2 need url.txt'
        if py_2.exists():
            os.remove(py_2)
        
        for sheet in [1,2]:
            file_path = f"{self.desktop_path / 'web spider' / f'sheet {sheet}_{self.last_Thrusday}--{self.this_Wednesday}.xlsx'}"
            if file_path.exists():
                os.remove(file_path)


    def check(self):
        self.files = []
        account = ""
        for root, dirs, temp_files in os.walk(self.desktop_path / 'web spider' / 'all of url'):
            self.files = temp_files
        print("请务必确保每一个公众号都打开并完整刷新")
        for acco in self.files:
            print(acco)
        input("!!完成后再按Enter")

    def fiddler_get(self):
        with open(f"{self.desktop_path / 'web spider' / 'py_2 need url.txt'}","r",encoding="utf-8") as f_allurl:
            allurl = f_allurl.readlines()
        for each in allurl:
            each_url = each.strip("\n")
            each_biz = each_url.split("__biz=")[1].split("&")[0][0:-6] + "==" 
            self.py_2_biz_with_url[each_biz] = each_url

    def backbone(self):
        fakeid_biz = {}
        with open(f"{self.desktop_path / 'web spider' / '数据_公众号信息.json'}","r",encoding="utf-8") as f_fakeid_biz:
            fakeid_biz = load(f_fakeid_biz)

        surplus = len(self.files)

        for i in self.files:
            account = i.split("_")[0]
            surplus -= 1
            print(f"{account}开始收集,还剩{surplus}个公众号")

            
            list_info = []
            open_mid_with_title = {}
            open_title_with_time = {}
            with open(f"{self.desktop_path / 'web spider' / 'all of url' / f'{account}_url.txt'}","r",encoding="utf-8") as open_url:
                list_info = open_url.readlines()

            with open(f"{self.desktop_path / 'web spider' / 'mid_with_title' / f'{account}.json'}","r") as open_mid:
                open_mid_with_title = load(open_mid)
            with open(f"{self.desktop_path / 'web spider' / 'title_with_time' / f'{account}.json'}","r") as open_title:
                open_title_with_time = load(open_title)
            #后续bug_check
            if(len(list_info) != len(open_mid_with_title)):
                for i in range(0,10):
                    print('sheet 2标题出现错误, 请结束后修改')
            if(len(list_info) != len(open_title_with_time)):
                for i in range(0,10):
                    print('sheet 2时间出现错误, 请结束后修改')
            ############
            self.account_s1.append(account)
            seg_publish_num_s1 = len(list_info)
            seg_read_num_s1 = 0     
            seg_old_like_num_s1 = 0 
            dict_time = []
            seg_publish_date_num_s1 = 0 
            seg_first_read_num_s1 = 0

            print(f"共{len(list_info)}篇")
            url = "https://" + self.py_2_biz_with_url[fakeid_biz[account]]
            # 优化处理  
            # change_mid = ''
            # ii = 0

            for i in list_info:
                mid = i.split("&")[1].split("=")[1]
                idx = i.split("&")[2].split("=")[1]
                sn = i.split("&")[3].split("=")[1]
                _biz = i.split("&")[0].split("?")[1].split("_biz=")[1]

                #优化处理
                # if(mid == change_mid):
                #     pass
                # else:
                #     ii = 0
                # ii += 1
                mid = mid + f'_{idx}'

                headers = {
                    "User-Agent": self.agent,
                }
                data = {
                    "mid": mid,
                    "idx": idx,
                    "sn": sn,
                    "__biz": _biz,
                    "is_only_read": "1",
                    "is_temp_url": "0",
                    "appmsg_type": "9",
                    "reward_uin_count":"0",
                }

                result = post(url,headers=headers,data=data,verify=False)
                try:
                    read_num = result.json()["appmsgstat"]["read_num"]
                except KeyError:
                    print("请求过于频繁,等待继续")
                    suspend_time = 5000
                    for suspend in range(suspend_time, -1, -1):
                        sleep(1)
                        print("等待中...{:>4}秒".format(suspend), end="\r")
                    print("\r")

                read_num = result.json()["appmsgstat"]["read_num"]
                old_like_num = result.json()["appmsgstat"]["old_like_num"]
                mid_title = open_mid_with_title[mid]
                title_time = open_title_with_time[open_mid_with_title[mid]][5:]

                seg_read_num_s1 += read_num
                seg_old_like_num_s1 += old_like_num
                if(title_time in dict_time):
                    pass
                else:
                    dict_time.append(title_time)
                    seg_publish_date_num_s1 += 1
                    seg_first_read_num_s1 += read_num

                self.all_data.write(f"{account}:--:{open_mid_with_title[mid]}:--:{open_title_with_time[open_mid_with_title[mid]][5:]}:--:{read_num}:--:{old_like_num}\n")

                if(read_num >= 1000):
                    self.index_sheet_2_num += 1
                    self.account_1000.append(account)
                    self.title_1000.append(mid_title)
                    self.read_num_1000.append(read_num)
                    self.old_like_num_1000.append(old_like_num)
                    title_time = round(int(title_time.split("-")[0]) + int(title_time.split("-")[1])/100, 2)
                    self.time_1000.append(title_time)

                sleep_time = 5 
                for remaining_time in range(sleep_time, -1, -1):
                    sleep(1)
                    print(f"等待爬取下一篇文章中...{remaining_time}秒", end="\r")
                print("\r")

            self.publish_num_s1.append(seg_publish_num_s1)
            self.read_num_s1.append(seg_read_num_s1)
            self.old_like_num_s1.append(seg_old_like_num_s1)
            self.first_read_num_s1.append(seg_first_read_num_s1)
            self.publish_date_num_s1.append(seg_publish_date_num_s1)

            print("\n")

            sleep_time = 10 
            for remaining_time in range(sleep_time, -1, -1):
                sleep(1)
                print("等待爬取下一公众号中...{:>2}秒".format(remaining_time), end="\r")
            print("\r")

        self.all_data.close()

    def write_in(self):
        file_path = []
        for sheet in [1,2]:
            file_path.append(f"{self.desktop_path / 'web spider' / f'sheet {sheet}_{self.last_Thrusday}--{self.this_Wednesday}.xlsx'}")
        
        for index_sheet_2 in range(1,self.index_sheet_2_num+1):
            self.list_index_1000.append(index_sheet_2)

        pd_account_s2 = pd.Series(self.account_1000, index=self.list_index_1000, name="公众号")
        pd_title_s2 = pd.Series(self.title_1000, index=self.list_index_1000, name="文章")
        pd_time_s2 = pd.Series(self.time_1000, index=self.list_index_1000, name="发布日期")
        pd_read_num_s2 = pd.Series(self.read_num_1000, index=self.list_index_1000, name="阅读量")
        pd_old_like_num_s2 = pd.Series(self.old_like_num_1000, index=self.list_index_1000, name="点赞量")
        pd_dict_s2 = {
            pd_account_s2.name: pd_account_s2,
            pd_title_s2.name: pd_title_s2,
            pd_time_s2.name: pd_time_s2,
            pd_read_num_s2.name: pd_read_num_s2,
            pd_old_like_num_s2.name: pd_old_like_num_s2,
        }
        df = pd.DataFrame(pd_dict_s2)
        df = df.set_index("公众号")
        df.to_excel(file_path[1])

        with open(f"{self.desktop_path / 'web spider' / '数据_公众号and类别.json'}","r",encoding="utf-8") as file:
            all_account = load(file)

        n = 0
        for acc in list(all_account.keys()):
            if acc not in self.account_s1:
                self.account_s1.append(acc)
                n += 1

        for acc in self.account_s1:
            self.category.append(all_account[acc])

        for index_sheet_1 in range(1,len(self.account_s1)+1):
            self.list_index_s1.append(index_sheet_1)

        publish_date_num_s1 += [0]*n
        publish_num_s1 += [0]*n
        read_num_s1 += [0]*n
        first_read_num_s1 += [0]*n
        old_like_num_s1 += [0]*n

        pd_account_s1 = pd.Series(self.account_s1, index=self.list_index_s1, name="公众号")
        pd_date_num_s1 = pd.Series(publish_date_num_s1, index=self.list_index_s1, name="发布次数").fillna(0)
        pd_publish_num_s1 = pd.Series(publish_num_s1, index=self.list_index_s1, name="发布篇数").fillna(0)
        pd_read_num_s1 = pd.Series(read_num_s1, index=self.list_index_s1, name="阅读总量").fillna(0)
        pd_first_read_num_s1 = pd.Series(first_read_num_s1, index=self.list_index_s1, name="头条阅读量").fillna(0)
        pd_old_like_num_s1 = pd.Series(old_like_num_s1, index=self.list_index_s1, name="点赞量").fillna(0)
        pd_category = pd.Series(self.category, index=self.list_index_s1, name="类别")
        pd_dict_s1 = {
            pd_account_s1.name: pd_account_s1,
            pd_date_num_s1.name: pd_date_num_s1,
            pd_publish_num_s1.name: pd_publish_num_s1,
            pd_read_num_s1.name: pd_read_num_s1,
            pd_first_read_num_s1.name: pd_first_read_num_s1,
            pd_old_like_num_s1.name: pd_old_like_num_s1,
            pd_category.name: pd_category,
        }
        df = pd.DataFrame(pd_dict_s1)
        df = df.set_index("公众号")
        df.to_excel(file_path[0])

        background_color = ""
        word_style = '华文细黑'
        size = 11

        for son_file in file_path:
            workbook = op.load_workbook(son_file)

            sheet = workbook.active
            header_row = sheet[1]
            for cell in header_row:
                if(son_file == file_path[0]):
                    background_color = "F7CAAC"
                else:
                    background_color = "A8D08E"
                cell.fill = op.styles.PatternFill(fill_type = 'solid', fgColor = background_color)

            font = op.styles.Font(name = word_style, size = size, color="404040")
            alignment_center = op.styles.Alignment(horizontal="center", vertical="center")
            alignment_left = op.styles.Alignment(horizontal="left", vertical="center")
            
            if(son_file == file_path[0]):
                sheet_n = 1
            else:
                sheet_n = 2

            col_n = 0
            for col in sheet.iter_cols():
                row_n = 0
                for cell in col:
                
                    if(col_n == 0 and row_n != 0):
                        cell.alignment = alignment_left
                    else:
                        if(sheet_n == 2 and col_n == 1 and row_n != 0):
                            cell.alignment = alignment_left
                        else:
                            cell.alignment = alignment_center
            
                    cell.font = font

                    row_n += 1
                col_n += 1
            
            workbook.save(son_file)
            workbook.close()



#-------------------------------------------------------
def main_1(desktop_path,last_Thrusday, this_Wednesday):
    create_1 = Create_1(desktop_path,last_Thrusday, this_Wednesday)
    create_1.first_check_file_and_dir()
    create_1.time_limit()
    create_1.drive()
    agent = create_1.get_url()
    create_1.backbone()
    return agent

def main_2(desktop_path,agent,last_Thrusday, this_Wednesday):
    create_2 = Create_2(desktop_path,agent,last_Thrusday, this_Wednesday)
    create_2.first_check_file_and_dir()
    create_2.check()
    create_2.fiddler_get()
    create_2.backbone()
    create_2.write_in()

def main():
    filterwarnings("ignore")

    init = Init()
    desktop_path = init.desktop_path
    init.start()
    last_Thrusday, this_Wednesday = init.time_restrict()

    agent = main_1(desktop_path,last_Thrusday, this_Wednesday)
    print('第一个文件已经完成,即将执行第二个文件.')
    os.system("pause")

    main_2(desktop_path,agent,last_Thrusday, this_Wednesday)
    os.system("pause")


if __name__=='__main__':
    main()